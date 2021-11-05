#! python3

#This is a tutorial on using python to read/write data into an excel sheet and perform
#analysis on it.

#In python, excel is handled using the "openpyxl" module.
import openpyxl

#Excel sheets (workbooks) use the ".xlsx" format.
#Columns are lettered (A,B,C...), rows are numbered (1,2,3...)
wb = openpyxl.load_workbook('myWorkbook.xlsx')
print(type(wb))

#Sheets/cells behave like dictionary objects where the key is the cell letter/number
#Keep in mind that indexing starts at 1, not 0.
print(wb.sheetnames)
sheet = wb['Sheet1']
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
print('Cell %s is %s' % (c.coordinate, c.value))
print(sheet['C1'].value)

#To swap column with number:
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1)) # Translate column 1 to a letter.
print(get_column_letter(900))
#The function column_index_from_string does the opposite.

#Get all cells
for row in sheet['A1':'C7']:
    for cell in row:
        print(cell.coordinate, cell.value)
    print('END OF ROW')

#Get cells in specified column
for cell in list(sheet.columns)[1]:
    print(cell.value)


#Creating a workbook and saving data.
wb = openpyxl.Workbook()  #Blank workbook
wb.sheetnames #Sheet is named ['Sheet']
sheet = wb.active
sheet.title = 'Spam Bacon Eggs'
wb.save('new_workbook.xlsx')

#Create/delete a new sheet
#wb.create_sheet()
#del wb.sheetnames[-1]

#Writing to cells
wb.create_sheet()
sheet = wb['Sheet']
sheet['A1'] = 'Hello world!'
sheet['B1'] = 5
wb.save('new_workbook.xlsx')


#Font style
from openpyxl.styles import Font
wb.create_sheet()
sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True) # Create a font.  #Can also set bold=True, name='Arial'
sheet['A1'].font = italic24Font # Apply the font to A1.
sheet['A1'] = 'Hello, world!'
wb.save('new_workbook.xlsx')


#Formulas are set like any other value in a cell.
sheet['A1']=5 ; sheet['A2']=10
sheet['A3']="=SUM(A1:A2)"

#Row height / column width  # =0 becomes hidden.
sheet.row_dimensions[1].height=70
sheet.column_dimensions['B'].width=20

#Merging/unmerging cells using .merge_cells() method
sheet.merge_cells('A1:A2')
sheet.unmerge_cells('A1:A2')

#Creating charts/graphs
for i in range(1, 11): # create some data in column A
    sheet['A' + str(i)] = i
refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.BarChart()
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')



wb.save('new_workbook.xlsx')

